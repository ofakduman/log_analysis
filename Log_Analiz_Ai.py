import os
import re
import logging
from collections import Counter
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from fpdf import FPDF, XPos, YPos
import streamlit as st
import matplotlib.pyplot as plt
from langchain_huggingface import HuggingFaceEndpoint
import io
from dotenv import load_dotenv


# Loglama ayarları
logging.basicConfig(filename='error_log_processing.log', level=logging.INFO, format='%(asctime)s - %(levelname)s: %(message)s')

load_dotenv()

# HuggingFaceEndpoint modelini başlatma
model = HuggingFaceEndpoint(
    repo_id="google/flan-t5-large",
    temperature=0.7, 
    max_length=100,
    huggingfacehub_api_token=os.environ["HUGGINGFACEHUB_API_TOKEN"]
)

# Error zaman dilimlerindeki sayıları hesaplayan fonksiyon
def count_errors_in_time_intervals(error_times):
    error_counts = {}
    for time in error_times:
        hour_minute_second = time.strftime('%H:%M:%S')
        if hour_minute_second not in error_counts:
            error_counts[hour_minute_second] = 1
        else:
            error_counts[hour_minute_second] += 1
    return error_counts

# Zaman dilimlerine göre en sık rastlanan hataları hesaplayan fonksiyon
def most_frequent_errors_by_time(error_times, error_descriptions):
    error_time_description = {}
    for time, description in zip(error_times, error_descriptions):
        hour_minute_second = time.strftime('%H:%M:%S')
        if description not in error_time_description:
            error_time_description[description] = {hour_minute_second: 1}
        elif hour_minute_second not in error_time_description[description]:
            error_time_description[description][hour_minute_second] += 1
        else:
            error_time_description[description][hour_minute_second] += 1
    return error_time_description


# Belirli bir yapıdaki cümleleri sayan fonksiyon
def count_specific_structure_sentences(log_list):
    cleaned_logs = clean_logs(log_list)
    structure_counter = Counter()
    structure_pattern = re.compile(r'\b[A-Za-z\s]+\b')
    for log in cleaned_logs:
        structures = structure_pattern.findall(log)
        specific_structure = [structure for structure in structures if 5 <= len(structure.split()) <= 20]
        structure_counter.update(specific_structure)
    return structure_counter
# Log dosyalarını temizleyen fonksiyon
def clean_logs(log_list):
    cleaned_logs = []
    for log in log_list:
        log = re.sub(r'\[WEBGATE\]', '', log)
        log = re.sub(r'at [\s\S]*', '', log)
        log = re.sub(r'SessionID:.*', '', log)
        log = re.sub(r'\{0\}', '', log)
        log = re.sub(r'\[.*?\]', '', log)
        log = re.sub(r'Risk durumu: \d+:', '', log)
        log = re.sub(r'Satılabilir stok : \d+,\d+', '', log)
        log = re.sub(r'Portfoyno: \d+', '', log)
        log = re.sub(r'\[TimesTen\]\[\w+\s\d+\.\d+\.\d+\.\d+ ODBC Driver\]\[TimesTen\]', '', log)
        log = re.sub(r'java.sql.SQLException: .*', '', log)
        log = re.sub(r'EYSException :.*', '', log)
        log = re.sub(r'at .* \(.*\)\s*', '', log)
        log = re.sub(r'Kullanici kodu: \w+', '', log)
        log = re.sub(r'EmirNo: \d+', '', log)
        
        cleaned_logs.append(log.strip())
    return cleaned_logs

# Loglardan hata mesajlarını çıkaran fonksiyon
def extract_errors_from_log(log):
    # Hata mesajlarını tespit etmek için regex deseni
    error_pattern = re.compile(r'ERROR:.*|Exception:.*', re.IGNORECASE)
    errors = error_pattern.findall(log)
    return errors


def generate_analysis_pdf(analyses):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)

    pdf.cell(200, 10, text="Excel Analysis Report", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')

    for sheet_name, analysis in analyses.items():
        pdf.cell(200, 10, text=f"Sheet: {sheet_name}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.multi_cell(200, 10, text=analysis)

    return bytes(pdf.output())


def analyze_excel(file_path):
    # Excel dosyasını oku
    df = pd.read_excel(file_path, sheet_name=None)

    # Her sayfayı incele ve modelden analiz iste
    analyses = {}
    for sheet_name, data in df.items():
        summary = data.describe(include='all')
        if len(summary) < 200:
            analysis_text = model.invoke(f"Summarize the following data analysis for the sheet '{sheet_name}':\n{summary}")
            analyses[sheet_name] = analysis_text

    return analyses




# Hatalara yönelik çözüm önerileri üretme fonksiyonu
# Hatalara yönelik çözüm önerileri üretme fonksiyonu
def generate_solutions(cleaned_logs):
    solutions = {}
    for log in cleaned_logs:
        errors = extract_errors_from_log(log)  # Logdan hataları çıkar
        for error in errors:
            # Hataları küçük parçalara ayırarak gönder
            split_errors = [error[i:i+100] for i in range(0, len(error), 100)]
            for split_error in split_errors:
                result = model.invoke(split_error)  # Modelin çözüm önerisi üretmesi
                solutions[split_error] = result
    return solutions



# PDF raporu oluşturma fonksiyonu
def generate_pdf_report(structure_sentences, solutions, categorized_errors):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)  # Arial yerine Helvetica

    pdf.cell(200, 10, text="Error Log Analysis Report", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    
    pdf.cell(200, 10, text="Specific Structure Sentences:", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    for sentence, count in structure_sentences.items():
        pdf.cell(200, 10, text=f"{sentence}: {count}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    
    pdf.cell(200, 10, text="Solutions:", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    for error, solution in solutions.items():
        pdf.cell(200, 10, text=f"{error}: {solution}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.cell(200, 10, text="Error Categories:", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    for category, count in categorized_errors.items():
        pdf.cell(200, 10, text=f"{category}: {count}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    
    return bytes(pdf.output())  # PDF'yi bytes olarak döndür




# Excel dosyasını oluşturma fonksiyonu
def generate_excel_report(logs, error_times, error_descriptions):
    output = io.BytesIO()
    workbook = Workbook()

    # Sayfa 1: Error Logs
    sheet1 = workbook.active
    sheet1.title = "Error Logs"

    # Başlıklar
    sheet1.append(["Tarih-Saat", "Hata Açıklaması"])

    for log in logs:
        if isinstance(log, dict):
            sheet1.append([log['Tarih-saat'], log['Error açıklaması']])
        else:
            # Handle log if it's a string, perhaps split by a delimiter
            log_parts = log.split(' ')  # Adjust the delimiter as needed
            sheet1.append([log_parts[0], ' '.join(log_parts[1:])])

    # Sayfa 2: Error Counts by Time
    sheet2 = workbook.create_sheet(title="Error Counts by Time")
    sheet2.append(["Tarih-Saat", "Hata Sayısı"])

    error_counts = count_errors_in_time_intervals(error_times)
    for time, count in error_counts.items():
        sheet2.append([time, count])

    # Sayfa 3: Frequent Errors by Time
    sheet3 = workbook.create_sheet(title="Frequent Errors by Time")
    sheet3.append(["Hata Açıklaması", "Tarih-Saat", "Hata Sayısı"])

    frequent_errors = most_frequent_errors_by_time(error_times, error_descriptions)
    for description, times in frequent_errors.items():
        for time, count in times.items():
            sheet3.append([description, time, count])

    workbook.save(output)
    output.seek(0)
    return output

# Hataları kategorilere göre sınıflandıran fonksiyon
def categorize_errors(error_descriptions):
    categories = {
        'Database': re.compile(r'database|sql|exception|timeout|connection', re.IGNORECASE),
        'Network': re.compile(r'network|timeout|connection|unreachable|disconnected', re.IGNORECASE),
        'Application': re.compile(r'nullpointer|illegalargument|runtime|application|app', re.IGNORECASE),
        'File System': re.compile(r'file|directory|not found|access denied|permission', re.IGNORECASE),
        'Security': re.compile(r'security|unauthorized|authentication|authorization|forbidden|login', re.IGNORECASE),
        'Performance': re.compile(r'performance|slow|timeout|delay|latency', re.IGNORECASE),
        'Memory': re.compile(r'memory|outofmemory|heap|stack', re.IGNORECASE),
        'Syntax': re.compile(r'syntax|parse|unexpected|unexpected token', re.IGNORECASE),
        'Configuration': re.compile(r'configuration|config|setting|parameter|missing|invalid', re.IGNORECASE),
        'Dependency': re.compile(r'dependency|library|module|package|not found', re.IGNORECASE),
        'Service': re.compile(r'service|unavailable|down|restart|failed', re.IGNORECASE),
        'Other': re.compile(r'.*')
    }

    categorized_errors = {category: 0 for category in categories}

    for description in error_descriptions:
        matched = False
        for category, pattern in categories.items():
            if pattern.search(description):
                categorized_errors[category] += 1
                matched = True
                break
        if not matched:
            categorized_errors['Other'] += 1

    return categorized_errors

# Streamlit arayüzü
st.title("Error Log Analysis Tool")

uploaded_files = st.file_uploader("Log dosyalarını yükleyin", accept_multiple_files=True)



if st.button("Logları İşle"):
    if uploaded_files:
        # Log işleme ve analiz kodları burada
        logs = []
        error_logs = []
        error_times = []
        error_descriptions = []
        illegal_chars = r'[^\w\s]'

        for file in uploaded_files:
            try:
                lines = file.read().decode('cp1254').splitlines()

                i = 0
                while i < len(lines):
                    if "ERROR" in lines[i]:
                        try:
                            error_start_time = re.search(r'\d{4}-\d{2}-\d{2}/\d{2}:\d{2}:\d{2}.\d{3}', lines[i]).group()
                            error_description = lines[i].split("ERROR", 1)[1].strip()
                            error_description = re.sub(r'\[[0-9]+\]', '', error_description)

                            i += 1
                            while i < len(lines) and not re.match(r'\d{4}-\d{2}-\d{2}/\d{2}:\d{2}:\d{2}.\d{3}', lines[i]):
                                error_description += "\n" + lines[i].strip()
                                i += 1

                            error_description = re.sub(illegal_chars, ' ', error_description)
                            error_description = error_description.replace('\x01', hex(ord('\x01')))

                            error_logs.append({
                                "Tarih-saat": pd.to_datetime(error_start_time),
                                "Error açıklaması": error_description,
                            })

                            error_times.append(pd.to_datetime(error_start_time))
                            error_descriptions.append(error_description)

                        except Exception as e:
                            logging.error(f"Error occurred while processing error logs: {str(e)}")
                    else:
                        i += 1

            except Exception as e:
                logging.error(f"Dosya okuma hatası: {str(e)}")
                st.error(f"Dosya okuma hatası: {str(e)}")
                st.stop()

        try:
            # Excel dosyasını oluştur ve bellek üzerinde sakla
            df = pd.DataFrame(error_logs)
            output = io.BytesIO()  # Bellekte dosya oluşturmak için
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Excel sayfalarını oluşturma ve yazma kodu
                df.to_excel(writer, index=False, sheet_name='Error Log')
                error_counts = df['Error açıklaması'].value_counts().reset_index()
                error_counts.columns = ['Error', 'Count']
                error_counts.to_excel(writer, index=False, sheet_name='Error Analiz')

                # Error Counts by Time Intervals sayfası
                error_counts_in_intervals = count_errors_in_time_intervals(error_times)
                sorted_error_counts = dict(sorted(error_counts_in_intervals.items(), key=lambda item: pd.to_datetime(item[0])))
                workbook = writer.book
                worksheet = workbook.create_sheet(title='Error Counts by Time Intervals')
                worksheet.append(['Zaman Aralığı', 'Hata Sayısı'])
                for time, count in sorted_error_counts.items():
                    worksheet.append([time, count])

                # Hata Sayısı Çubuğu Grafiği
                chart = BarChart()
                chart.title = "Hata Sayısı"
                chart.y_axis.title = "Hata Sayısı"
                chart.x_axis.title = "Zaman Aralığı"
                data = Reference(worksheet, min_col=2, min_row=2, max_row=len(sorted_error_counts) + 1)
                cats = Reference(worksheet, min_col=1, min_row=2, max_row=len(sorted_error_counts) + 1)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.height = 10
                chart.width = 20
                worksheet.add_chart(chart, "D2")

                # Hata Zamanları Çizgi Grafiği
                try:
                    error_time_counts = Counter(error_times)
                    sorted_error_times = dict(sorted(error_time_counts.items()))

                    plt.figure(figsize=(12, 6))
                    plt.plot(list(sorted_error_times.keys()), list(sorted_error_times.values()), marker='o')
                    plt.title("Hata Sayıları (Zaman Aralıklarına Göre)")
                    plt.xlabel("Zaman")
                    plt.ylabel("Hata Sayısı")
                    plt.xticks(rotation=45)

                    plot_path = io.BytesIO()
                    plt.savefig(plot_path, format='png')
                    plt.close()
                    plot_path.seek(0)
                    img = Image(plot_path)
                    worksheet.add_image(img, 'A20')

                except Exception as e:
                    logging.error(f"Grafik oluşturulurken bir hata oluştu: {str(e)}")
                    st.error(f"Grafik oluşturulurken bir hata oluştu: {str(e)}")

                # Specific Structure Sentences sayfası
                specific_structure_repeated = count_specific_structure_sentences(df['Error açıklaması'])
                if specific_structure_repeated:
                    structure_sheet = workbook.create_sheet(title='Specific Structure Sentences')
                    structure_sheet.append(['Cümle Yapısı', 'Sayısı'])
                    for structure, count in specific_structure_repeated.items():
                        structure_sheet.append([structure, count])

                # Error Categories sayfası
                categorized_errors = categorize_errors(df['Error açıklaması'])
                if categorized_errors:
                    category_sheet = workbook.create_sheet(title='Error Categories')
                    category_sheet.append(['Kategori', 'Hata Sayısı'])
                    for category, count in categorized_errors.items():
                        category_sheet.append([category, count])

            # Excel dosyasını PDF ile analiz için kapatma
            output.seek(0)

            # Excel analizini yap ve PDF oluştur
            analyses = analyze_excel(output)
            pdf_data = generate_analysis_pdf(analyses)

            # Streamlit'te dosya indirme butonları ekleme
            st.download_button("Excel Raporunu İndir", data=output, file_name="error_logs.xlsx")
            st.download_button("Analiz PDF Raporunu İndir", data=pdf_data, file_name="excel_analysis_report.pdf")

        except Exception as e:
            logging.error(f"Loglar işlenirken hata oluştu: {str(e)}")
            st.error(f"Loglar işlenirken bir hata oluştu: {str(e)}")
            st.stop()
    else:
        st.warning("Lütfen en az bir log dosyası yükleyin.")

